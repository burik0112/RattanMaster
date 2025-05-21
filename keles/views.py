from collections import defaultdict
from datetime import datetime

import openpyxl
from django.db.models import Count, Sum, Q
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils.formats import date_format
from openpyxl.styles import Side, Border, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from admiin.decorators import role_required
from index.models import TransferToInventory, CategoryModel
from keles.models import InvoiceCreateKeles, ProductEntryKeles, RemaingInventoryKeles, InvoiceKeles


# Create your views here.


def Dashboard(request):
    total_quantity = InvoiceCreateKeles.objects.aggregate(Sum('quantity'))['quantity__sum']
    total_product_in = ProductEntryKeles.objects.aggregate(Sum('quantity'))['quantity__sum']
    total_remaing = RemaingInventoryKeles.objects.aggregate(Sum('quantity'))['quantity__sum']
    # Если нет данных, установим значение в 0
    if total_quantity is None:
        total_quantity = 0

    if total_product_in is None:
        total_product_in = 0

    if total_remaing is None:
        total_product_in = 0

    # Аннотируем категории, добавляя количество продуктов в каждой категории
    categories_with_product_count = CategoryModel.objects.annotate(
        count=Count('products')
    )
    transfer_to = TransferToInventory.objects.annotate(
        count=Count('transfer_to')
    )

    # Возвращаем результат в шаблон
    return render(request, 'keles/catalog_category.html', {
        'query': categories_with_product_count,
        'total_quantity': total_quantity,
        'total_product_in': total_product_in,
        'total_remaing': total_remaing,
        'transfer_to': transfer_to
    })


def Base(request, pk):
    query = InvoiceCreateKeles.objects.filter(name__id=pk)
    return render(request, 'keles/index.html', {'query': query})


def Detail(request, pk):
    query = InvoiceCreateKeles.objects.filter(pk=pk)
    return render(request, 'keles/detail.html', {'query': query})

@role_required(['Оператор склада Келес', 'Начальник'])
def InvoiceCreateFromKeles(request):
    invoice = InvoiceCreateKeles.objects.all().order_by('-created_at')
    search_query = request.GET.get('search', '')  # Now this refers to the model

    product_id = request.GET.get('product_id')
    product_name = request.GET.get('product_name')
    product_size = request.GET.get('product_size')
    color = request.GET.get('product_color')
    transfer = request.GET.get('transfer')
    quantity = request.GET.get('quantity')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Apply filters based on the request
    if product_id:
        invoice = invoice.filter(name__id=product_id)
    if product_name:
        invoice = invoice.filter(name__title__icontains=product_name)
    if product_size:
        invoice = invoice.filter(size__title__icontains=product_size)
    if color:
        invoice = invoice.filter(color__title=color)
    if transfer:
        invoice = invoice.filter(product_to__title__icontains=transfer)
    if quantity:
        invoice = invoice.filter(quantity=quantity)

    # Apply date filter if provided
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            invoice = invoice.filter(created_at__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass  # Ignore invalid date formats

    if request.GET.get('export') == 'excel':
        return export_to_excel(invoice)

        return response
    return render(request, 'keles/invoice-catalog.html',
                  {'invoice': invoice,
                   'search_query': search_query})

@role_required(['Оператор склада Келес', 'Начальник'])
def export_to_excel(invoices):
    # Create a new workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    # Define the header row
    headers = [
        'Invoice ID', 'Product Name', 'Size', 'Color', 'Transfer To', 'Quantity', 'Created At'
    ]
    ws.append(headers)

    # Populate data rows from the filtered invoices
    for inv in invoices:
        row = [
            str(inv.id),
            inv.name.title if inv.name else 'N/A',
            inv.size.title if inv.size else 'N/A',
            inv.color.title if inv.color else 'N/A',
            inv.product_to.title if inv.product_to else 'N/A',
            str(inv.quantity),
            inv.created_at.strftime('%Y-%m-%d')  # Format the date
        ]
        ws.append(row)

    # Create an HTTP response with content-type set for Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="InvoiceData.xlsx"'

    # Save the workbook directly to the response object
    wb.save(response)

    return response

@role_required(['Сотрудник приемки Келес', 'Начальник'])
def ProductInKeles(request):
    product_in = ProductEntryKeles.objects.order_by('-created_at')
    product_id = request.GET.get('product_id')
    product_name = request.GET.get('product_name')
    product_size = request.GET.get('product_size')
    color = request.GET.get('color')
    transfer = request.GET.get('transfer')
    quantity = request.GET.get('quantity')
    created_at = request.GET.get('created_at')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Apply the filters based on the search parameters
    if product_id:
        product_in = product_in.filter(name__id=product_id)
    if product_name:
        product_in = product_in.filter(name__title__icontains=product_name)
    if product_size:
        product_in = product_in.filter(size__title__icontains=product_size)
    if color:
        product_in = product_in.filter(color__title__icontains=color)
    if transfer:
        product_in = product_in.filter(product_in__title__icontains=transfer)
    if quantity:
        product_in = product_in.filter(quantity=quantity)

    # Apply the date filter if the 'created_at' filter is provided
    if created_at:
        try:
            # Convert the 'created_at' string into a date object
            created_at_date = datetime.strptime(created_at, '%Y-%m-%d').date()
            product_in = product_in.filter(created_at=created_at_date)
        except ValueError:
            pass  # If the date format is invalid, simply ignore it

    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            product_in = product_in.filter(created_at__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass  # Ignore invalid date formats

    # If export is requested, generate Excel file
    if request.GET.get('export') == 'excel':
        return export_from_excel(product_in)

    return render(request, 'keles/product_in-catalog.html', {'product_in': product_in})

@role_required(['Сотрудник приемки Келес', 'Начальник'])
def export_from_excel(invoices):
    # Create a new workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    # Define the header row
    headers = [
        'Invoice ID', 'Product Name', 'Size', 'Color', 'TransferFrom', 'Quantity', 'Created At'
    ]
    ws.append(headers)

    # Populate data rows from the filtered invoices
    for inv in invoices:
        row = [
            str(inv.id),
            inv.name.title if inv.name else 'N/A',
            inv.size.title if inv.size else 'N/A',
            inv.color.title if inv.color else 'N/A',
            inv.product_in.title if inv.product_in else 'N/A',
            str(inv.quantity),
            inv.created_at.strftime('%Y-%m-%d')  # Format the date
        ]
        ws.append(row)

    # Create an HTTP response with content-type set for Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="InvoiceData.xlsx"'

    # Save the workbook directly to the response object
    wb.save(response)

    return response

@role_required(['Сотрудник приемки Келес', 'Начальник'])
def RemaingListKeles(request):
    remaing = RemaingInventoryKeles.objects.all().order_by('-name')
    return render(request, 'keles/remaing-list.html', {'remaing': remaing})

@role_required(['Менеджер склада Келес', 'Начальник'])
def TurnoverKeles(request):
    search_name = request.GET.get('search_name', '').strip().lower()
    search_id = request.GET.get('search_id', '').strip().lower()
    search_size = request.GET.get('search_size', '').strip().lower()
    search_color = request.GET.get('search_color', '').strip().lower()

    # Preload all product info (to avoid querying in loop)
    all_entries = list(ProductEntryKeles.objects.select_related('name', 'size', 'color').all())
    all_remaining = list(RemaingInventoryKeles.objects.select_related('name', 'size', 'color').all())
    all_invoice = list(InvoiceCreateKeles.objects.select_related('name', 'size', 'color').all())

    # Aggregate quantities by (name_id, size_id, color_id)
    def aggregate_by(items):
        result = defaultdict(int)
        for obj in items:
            key = (obj.name_id, obj.size_id, obj.color_id)
            result[key] += obj.quantity
        return result

    entry_totals = aggregate_by(all_entries)
    remaining_totals = aggregate_by(all_remaining)
    invoice_totals = aggregate_by(all_invoice)

    # Combine all keys
    all_keys = set(entry_totals) | set(remaining_totals) | set(invoice_totals)

    inventory_data = []

    # Build a single reference map for title resolution
    ref_objects = {}
    for obj in all_entries + all_remaining + all_invoice:
        key = (obj.name_id, obj.size_id, obj.color_id)
        if key not in ref_objects:
            ref_objects[key] = obj

    for key in all_keys:
        name_id, size_id, color_id = key
        total_in = entry_totals.get(key, 0)
        total_remaining = remaining_totals.get(key, 0)
        total_invoice = invoice_totals.get(key, 0)
        remaining_stock = total_in + total_remaining - total_invoice

        obj = ref_objects.get(key)
        if not obj:
            continue

        category_title = obj.name.title.lower() if obj.name else ''
        size_title = obj.size.title.lower() if obj.size else ''
        color_title = obj.color.title.lower() if obj.color else ''
        category_id = str(obj.name.id) if obj.name else ''

        if (
                search_name in category_title and
                search_id in category_id and
                search_size in size_title and
                search_color in color_title
        ):
            inventory_data.append({
                'category_id': obj.name.id,
                'category_title': obj.name.title,
                'size_title': obj.size.title,
                'color_title': obj.color.title,
                'total_in': total_in,
                'total_remaining': total_remaining,
                'total_invoice': total_invoice,
                'remaining_stock': remaining_stock,
            })

    inventory_data.sort(key=lambda x: (
        x['category_title'].lower(),
        x['size_title'].lower(),
        x['color_title'].lower()
    ))

    # Excel download
    if 'download_excel' in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"

        headers = ['ID', 'Category', 'Size', 'Color', 'Remaining']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header

        for row_num, data in enumerate(inventory_data, 2):
            ws[f'A{row_num}'] = data['category_id']
            ws[f'B{row_num}'] = data['category_title']
            ws[f'C{row_num}'] = data['size_title']
            ws[f'D{row_num}'] = data['color_title']
            ws[f'E{row_num}'] = data['remaining_stock']

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Turnover.xlsx"'
        wb.save(response)
        return response

    return render(request, 'keles/turnover.html', {'inventory_data': inventory_data})


def Client_ReportKeles(request, pk):
    # Fetching the data from the database
    query = InvoiceCreateKeles.objects.filter(product_to__id=pk)

    if request.GET.get('export') == 'excel':  # Check if the export parameter is set
        # Create a workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Data"

        # Define the header row (you can customize this as per your model fields)
        headers = ['ID', 'Product Name', 'Color', 'Size', 'Product To', 'Quantity',
                   'Created_at']  # Replace with your model fields
        ws.append(headers)

        # Populate the data rows from the queryset
        for invoice in query:
            row = [
                str(invoice.id),  # Invoice ID
                str(invoice.name) if invoice.name else 'N/A',  # CategoryModel name (using .name)
                str(invoice.color) if invoice.color else 'N/A',  # ColorModel color (using .color)
                str(invoice.size) if invoice.size else 'N/A',
                str(invoice.product_to.title) if invoice.product_to else 'N/A',
                # TransferToInventory name (using .name)
                str(invoice.quantity),  # Quantity
                invoice.created_at.strftime('%Y-%m-%d'),  # Date formatted as YYYY-MM-DD
            ]
            row = [str(cell) for cell in row]
            ws.append(row)

            # Create an HTTP response with content-type set for Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Client List.xlsx'

        # Save the workbook directly to the response object
        wb.save(response)

        return response

        # If no export requested, render the normal dashboard view
    return render(request, 'keles/dashboard.html', {'query': query})
    # SizeModel size (using .size)



def invoice_list(request):
    # Извлекаем все накладные
    invoices = InvoiceKeles.objects.all().order_by('-created_at')  # Сортируем по дате создания (по убыванию)

    # Передаем данные в контекст
    context = {
        'invoices': invoices
    }

    return render(request, 'keles/invoice_list.html', context)



def invoice_detail(request, pk):
    invoice = get_object_or_404(InvoiceKeles, pk=pk)
    items = invoice.items.all()  # через related_name='items'

    context = {
        'invoice': invoice,
        'items': items
    }
    return render(request, 'keles/invoice_detail.html', context)



def export_to_excel(request, invoice_id):
    try:
        invoice = InvoiceKeles.objects.get(pk=invoice_id)
    except InvoiceKeles.DoesNotExist:
        return HttpResponse("Накладная не найдена.")

    items = invoice.items.select_related('name', 'size', 'color')
    if not items.exists():
        return HttpResponse("Нет товаров для экспорта.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Накладная"

    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    invoice_date = date_format(invoice.created_at, 'd.m.Y')

    # Шапка
    ws.merge_cells('A1:F1')
    ws['A1'] = f"НАКЛАДНАЯ № {invoice.number}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center

    ws.merge_cells('A2:F2')
    ws['A2'] = f'от "{invoice_date}"'
    ws['A2'].alignment = center

    ws['A4'] = "От кого:"
    ws['B4'] = "OOO RATTAN MASTER"
    ws['A5'] = "Кому:"
    ws['B5'] = invoice.product_to.title if invoice.product_to else "—"
    ws['D5'] = "Через________________"

    # Таблица
    headers = ["№ п/п", "Наименование", "Размер", "Цвет", "Количество"]
    ws.append([])
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 20 if col_num != 1 else 8

    row = 8
    total_quantity = 0

    for index, item in enumerate(items, start=1):
        values = [
            index,
            item.name.title,
            item.size.title if item.size else '',
            item.color.title if item.color else '',
            f"{item.quantity} шт"
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = center
            cell.border = border
        total_quantity += item.quantity
        row += 1

    # Итого
    total_row = ["", "", "", "Итого:", f"{total_quantity} шт"]
    for col, val in enumerate(total_row, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = bold_font
        cell.alignment = center
        cell.border = border
    row += 2

    # Подписи
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Сдал: ________________   Ф. И. О."
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = "Принял: ________________   Ф. И. О."

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Invoice_{invoice.number}.xlsx"'
    wb.save(response)
    return response