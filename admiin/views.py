import openpyxl
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.formats import date_format
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from admiin.decorators import role_required
from admiin.forms import CategoryForm, SizeForm, ColorForm, InvoiceCreateForm, ProductInCreateForm, RemaingCreateForm, \
    InvoiceCreateKelesForm, ProductInCreateKelesForm, RemaingCreateKelesForm, InvoiceKelesForm, InvoiceXasanboyForm
from index.models import CategoryModel, SizeModel, ColorModel, RemaingInventoryModel, InvoiceCreateModel, \
    TransferToInventory, TransferFromInventory, ProductEntry, Invoice
from keles.models import RemaingInventoryKeles, InvoiceCreateKeles, ProductEntryKeles, InvoiceKeles


@role_required([])
def AdminCategoryEdit(request, pk):
    category = get_object_or_404(CategoryModel, pk=pk)

    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('pages:category')
    else:
        form = CategoryForm(instance=category)

    return render(request, 'edit/category-edit.html', {'form': form})


@role_required([])
def CategoryDelete(request, pk):
    query = get_object_or_404(CategoryModel, pk=pk)
    if request:
        query.delete()
        return redirect('pages:category')
    return render(request, 'category-cat.html')


@role_required([])
def AdminSizeEdit(request, pk):
    size = get_object_or_404(SizeModel, pk=pk)

    if request.method == 'POST':
        form = SizeForm(request.POST, instance=size)
        if form.is_valid():
            form.save()
            return redirect('pages:size')
    else:
        form = SizeForm(instance=size)

    return render(request, 'edit/size-edit.html', {'form': form})


@role_required([])
def SizeDelete(request, pk):
    query = get_object_or_404(SizeModel, pk=pk)
    if request:
        query.delete()
        return redirect('pages:size')
    return render(request, 'size-catalog.html')


@role_required([])
def Color_create_or_edit(request, pk=None):
    if pk:
        color = get_object_or_404(ColorModel, pk=pk)
    else:
        color = None

    if request.method == 'POST':
        form = ColorForm(request.POST, instance=color)
        if form.is_valid():
            form.save()
            return redirect('pages:color')
    else:
        form = ColorForm(instance=color)

    return render(request, 'file/color-add.html', {'form': form})


@role_required([])
def ColorDelete(request, pk):
    query = get_object_or_404(ColorModel, pk=pk)
    if request:
        query.delete()
        return redirect('pages:color')
    return render(request, 'size-catalog.html')


@role_required([])
def Category_create_or_edit(request, pk=None):
    if pk:
        cat = get_object_or_404(CategoryModel, pk=pk)
    else:
        cat = None

    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=cat)
        if form.is_valid():
            form.save()
            return redirect('pages:category')
    else:
        form = CategoryForm(instance=cat)

    return render(request, 'file/category-add.html', {'form': form})


@role_required([])
def Size_create_or_edit(request, pk=None):
    if pk:
        size = get_object_or_404(SizeModel, pk=pk)
    else:
        size = None

    if request.method == 'POST':
        form = SizeForm(request.POST, instance=size)
        if form.is_valid():
            form.save()
            return redirect('pages:size')
    else:
        form = SizeForm(instance=size)

    return render(request, 'file/size-add.html', {'form': form})


@role_required(['Сотрудник приемки', 'Начальник'])
def AddRemaing(request):
    if request.method == 'POST':
        name_id = request.POST.get('name')
        size_id = request.POST.get('size')
        color_id = request.POST.get('color')
        quantity = request.POST.get('quantity')

        name = CategoryModel.objects.get(id=name_id)
        size = SizeModel.objects.get(id=size_id)
        color = ColorModel.objects.get(id=color_id)

        RemaingInventoryModel.objects.create(
            name=name,
            size=size,
            color=color,
            quantity=quantity
        )

        return redirect('pages:remaing_list')

    # For GET: show the form and pass dropdown data
    categories = CategoryModel.objects.all()
    sizes = SizeModel.objects.all()
    colors = ColorModel.objects.all()

    return render(request, 'file/remaing-add.html', {
        'categories': categories,
        'sizes': sizes,
        'colors': colors,
    })


@role_required(['Оператор склада', 'Начальник'])
def InvoiceCreate(request):
    if request.method == 'POST':
        name_ids = request.POST.getlist('name')
        size_ids = request.POST.getlist('size')
        color_ids = request.POST.getlist('color')
        quantities = request.POST.getlist('quantity')
        product_to_id = request.POST.get('product_to')

        # Создаём накладную
        product_to = TransferToInventory.objects.get(id=product_to_id)
        invoice = Invoice.objects.create(product_to=product_to)

        # Добавляем продукты
        items = []
        for name_id, size_id, color_id, quantity in zip(name_ids, size_ids, color_ids, quantities):
            item = InvoiceCreateModel.objects.create(
                invoice=invoice,
                name=CategoryModel.objects.get(id=name_id),
                size=SizeModel.objects.get(id=size_id),
                color=ColorModel.objects.get(id=color_id) if color_id else None,
                quantity=int(quantity)
            )
            items.append(item.id)  # Сохраняем ID товаров

        # Сохраняем ID в сессии для экспорта
        request.session['last_invoice_ids'] = items

        # Перенаправляем на страницу для экспорта
        return redirect('add:export_to_excel')

    context = {
        'categories': CategoryModel.objects.all(),
        'sizes': SizeModel.objects.all(),
        'colors': ColorModel.objects.all(),
        'transfer_to': TransferToInventory.objects.all()
    }
    return render(request, 'file/product_add.html', context)


@role_required(['Сотрудник приемки', 'Начальник'])
def EntryCreate(request):
    if request.method == 'POST':
        name_id = request.POST.get('name')
        size_id = request.POST.get('size')
        color_id = request.POST.get('color')
        product_in_id = request.POST.get('product_in')
        quantity = request.POST.get('quantity')

        name = CategoryModel.objects.get(id=name_id)
        size = SizeModel.objects.get(id=size_id)
        color = ColorModel.objects.get(id=color_id)
        product_in = TransferFromInventory.objects.get(id=product_in_id)

        created_at = timezone.now().date()

        ProductEntry.objects.create(
            name=name,
            size=size,
            color=color,
            product_in=product_in,
            quantity=quantity,
            created_at=created_at
        )

        return redirect('pages:product_in-list')  # Change this to your actual URL name

    # For GET request: Show the form and pass dropdown data
    categories = CategoryModel.objects.all()
    sizes = SizeModel.objects.all()
    colors = ColorModel.objects.all()
    product_in = TransferFromInventory.objects.all()

    return render(request, 'file/product-in.html', {
        'categories': categories,
        'sizes': sizes,
        'colors': colors,
        'product_in': product_in,
    })


@role_required(['Оператор склада', 'Начальник'])
def InvoiceEdit(request, pk):
    obj = get_object_or_404(InvoiceCreateModel, pk=pk)
    if request.method == 'POST':
        form = InvoiceCreateForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('pages:invoice-list')  # replace with your actual detail or success view
    else:
        form = InvoiceCreateForm(instance=obj)
    return render(request, 'edit/product-in-edit.html', {'form': form})


@role_required(['Оператор склада', 'Начальник'])
def InvoiceDelete(request, pk):
    invoice = get_object_or_404(InvoiceCreateModel, pk=pk)
    invoice.delete()
    query_string = request.GET.urlencode()
    return redirect(f"{reverse('pages:invoice-list')}?{query_string}")


@role_required(['Сотрудник приемки', 'Начальник'])
def ProductInEdit(request, pk):
    obj = get_object_or_404(ProductEntry, pk=pk)
    if request.method == 'POST':
        form = ProductInCreateForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('pages:product_in-list')  # replace with your actual detail or success view
    else:
        form = ProductInCreateForm(instance=obj)
    return render(request, 'edit/product-in-edit.html', {'form': form})


@role_required(['Сотрудник приемки', 'Начальник'])
def ProductInDelete(request, pk):
    query = get_object_or_404(ProductEntry, pk=pk)
    if request:
        query.delete()
        return redirect('pages:product_in-list')  # Change to your actual URL name
    return render(request, 'product_in-catalog.html')


@role_required(['Сотрудник приемки', 'Начальник'])
def RemaingEdit(request, pk):
    obj = get_object_or_404(RemaingInventoryModel, pk=pk)
    if request.method == 'POST':
        form = RemaingCreateForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('pages:remaing_list')  # replace with your actual detail or success view
    else:
        form = RemaingCreateForm(instance=obj)
    return render(request, 'edit/remaing-edit.html', {'form': form})


@role_required(['Сотрудник приемки'])
def RemaingDelete(request, pk):
    query = get_object_or_404(RemaingInventoryModel, pk=pk)
    if request:
        query.delete()
        return redirect('pages:remaing_list')  # Change to your actual URL name
    return render(request, 'remaing-list.html')


@role_required(['Оператор склада', 'Начальник'])
def export_to_excel(request):
    ids = request.session.get('last_invoice_ids', [])
    if not ids:
        return HttpResponse("Нет данных для экспорта.")

    # Получаем товары по ID
    items = InvoiceCreateModel.objects.filter(id__in=ids).select_related('invoice', 'name', 'size', 'color')

    if not items.exists():
        return HttpResponse("Нет товаров для экспорта.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Накладная"

    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    invoice_item = items.first()
    invoice = invoice_item.invoice
    invoice_date = date_format(invoice.created_at, 'd.m.Y')

    # Шапка
    ws.merge_cells('A1:F1')
    ws['A1'] = f"НАКЛАДНАЯ № {invoice.number}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center

    ws.merge_cells('A2:F2')
    ws['A2'] = f'от "{invoice_date}"'
    ws['A2'].alignment = center

    ws['A4'] = "От кого:"
    ws['B4'] = "OOO RATTAN MASTER"
    ws['A5'] = "Кому:"
    ws['B5'] = invoice.product_to.title if invoice.product_to else "—"
    ws['D5'] = "Через________________"

    # Таблица
    headers = ["№ п/п", "Наименование", "Размер", "Цвет", "Количество"]
    ws.append([])
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 20 if col_num != 1 else 8

    row = 8
    total_quantity = 0

    for index, item in enumerate(items, start=1):
        values = [
            index,
            item.name.title,
            item.size.title if item.size else '',
            item.color.title if item.color else '',
            f"{item.quantity} шт"
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = center
            cell.border = border
        total_quantity += item.quantity
        row += 1

    # Итого
    total_row = ["", "", "", "Итого:", f"{total_quantity} шт"]
    for col, val in enumerate(total_row, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = bold_font
        cell.alignment = center
        cell.border = border
    row += 2

    # Подписи
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Сдал: ________________   Ф. И. О."
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = "Принял: ________________   Ф. И. О."

    # Скачивание
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Invoice_{invoice.number}.xlsx"'
    wb.save(response)
    return response


# All code starting for keles from this collumb
@role_required(['Сотрудник приемки Келес', 'Начальник'])
def AddRemaingKeles(request):
    if request.method == 'POST':
        name_id = request.POST.get('name')
        size_id = request.POST.get('size')
        color_id = request.POST.get('color')
        quantity = request.POST.get('quantity')

        name = CategoryModel.objects.get(id=name_id)
        size = SizeModel.objects.get(id=size_id)
        color = ColorModel.objects.get(id=color_id)

        RemaingInventoryKeles.objects.create(
            name=name,
            size=size,
            color=color,
            quantity=quantity
        )

        return redirect('keles:remaing_list')

    categories = CategoryModel.objects.all()
    sizes = SizeModel.objects.all()
    colors = ColorModel.objects.all()

    return render(request, 'keles-add/remaing-add.html', {
        'categories': categories,
        'sizes': sizes,
        'colors': colors,
    })


@role_required(['Оператор склада Келес', 'Начальник'])
def InvoiceCreate2Model(request):
    if request.method == 'POST':
        name_ids = request.POST.getlist('name')
        size_ids = request.POST.getlist('size')
        color_ids = request.POST.getlist('color')
        quantities = request.POST.getlist('quantity')
        product_to_id = request.POST.get('product_to')

        # Создаём накладную
        product_to = TransferToInventory.objects.get(id=product_to_id)
        invoice = InvoiceKeles.objects.create(product_to=product_to)

        # Добавляем продукты
        items = []
        for name_id, size_id, color_id, quantity in zip(name_ids, size_ids, color_ids, quantities):
            item = InvoiceCreateKeles.objects.create(
                invoice=invoice,
                name=CategoryModel.objects.get(id=name_id),
                size=SizeModel.objects.get(id=size_id),
                color=ColorModel.objects.get(id=color_id) if color_id else None,
                product_to=TransferToInventory(id=name_id),
                quantity=int(quantity)
            )
            items.append(item.id)  # Сохраняем ID товаров

        # Сохраняем ID в сессии для экспорта
        request.session['last_invoice_ids'] = items

        # Перенаправляем на страницу для экспорта
        return redirect('add:export_to_excel')

    context = {
        'categories': CategoryModel.objects.all(),
        'sizes': SizeModel.objects.all(),
        'colors': ColorModel.objects.all(),
        'transfer_to': TransferToInventory.objects.all()
    }
    return render(request, 'keles-add/product_add.html', context)


@role_required(['Сотрудник приемки Келес', 'Начальник'])
def EntryCreateKeles(request):
    if request.method == 'POST':
        name_id = request.POST.get('name')
        size_id = request.POST.get('size')
        color_id = request.POST.get('color')
        product_in_id = request.POST.get('product_in')
        quantity = request.POST.get('quantity')

        name = CategoryModel.objects.get(id=name_id)
        size = SizeModel.objects.get(id=size_id)
        color = ColorModel.objects.get(id=color_id)
        product_in = TransferFromInventory.objects.get(id=product_in_id)

        created_at = timezone.now().date()

        ProductEntryKeles.objects.create(
            name=name,
            size=size,
            color=color,
            product_in=product_in,
            quantity=quantity,
            created_at=created_at
        )

        return redirect('keles:product_in-list')

    categories = CategoryModel.objects.all()
    sizes = SizeModel.objects.all()
    colors = ColorModel.objects.all()
    product_in = TransferFromInventory.objects.all()

    return render(request, 'keles-add/product-in.html', {
        'categories': categories,
        'sizes': sizes,
        'colors': colors,
        'product_in': product_in,
    })


@role_required(['Оператор склада', 'Начальник'])
def InvoiceEditKeles(request, pk):
    invoice = get_object_or_404(InvoiceCreateKeles, pk=pk)
    if request.method == 'POST':
        # handle form
        form = InvoiceCreateKelesForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            query_string = request.GET.urlencode()
            return redirect(f"{reverse('keles:invoice-list')}?{query_string}")
    else:
        form = InvoiceCreateKelesForm(instance=invoice)
    return render(request, 'keles-add/invoice-edit.html', {'form': form})


@role_required(['Оператор склада', 'Начальник'])
def InvoiceDeleteKeles(request, pk):
    invoice = get_object_or_404(InvoiceCreateKeles, pk=pk)
    invoice.delete()
    query_string = request.GET.urlencode()
    return redirect(f"{reverse('keles:invoice-list')}?{query_string}")


@role_required(['Сотрудник приемки Келес', 'Начальник'])
def ProductInEditKeles(request, pk):
    obj = get_object_or_404(ProductEntryKeles, pk=pk)
    if request.method == 'POST':
        form = ProductInCreateKelesForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('keles:product_in-list')  # replace with your actual detail or success view
    else:
        form = ProductInCreateKelesForm(instance=obj)
    return render(request, 'keles-add/product-in-edit.html', {'form': form})


@role_required(['Сотрудник приемки Келес', 'Начальник'])
def ProductInDeleteKeles(request, pk):
    query = get_object_or_404(ProductEntryKeles, pk=pk)
    if request:
        query.delete()
        return redirect('keles:product_in-list')  # Change to your actual URL name
    return render(request, 'keles/product_in-catalog.html')


@role_required(['Сотрудник приемки Келес', 'Начальник'])
def RemaingEditKeles(request, pk):
    obj = get_object_or_404(RemaingInventoryKeles, pk=pk)
    if request.method == 'POST':
        form = RemaingCreateKelesForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('keles:remaing_list')  # replace with your actual detail or success view
    else:
        form = RemaingCreateKelesForm(instance=obj)
    return render(request, 'keles-add/remaing-edit.html', {'form': form})


@role_required(['Сотрудник приемки Келес', 'Начальник'])
def RemaingDeleteKeles(request, pk):
    query = get_object_or_404(RemaingInventoryKeles, pk=pk)
    if request:
        query.delete()
        return redirect('keles:remaing_list')  # Change to your actual URL name
    return render(request, 'keles/remaing-list.html')


@role_required(['Оператор склада Келес', 'Начальник'])
def export_to_excelkeles(request):
    ids = request.session.get('last_invoice_ids', [])
    if not ids:
        return HttpResponse("Нет данных для экспорта.")

    items = InvoiceCreateKeles.objects.filter(id__in=ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Накладная"

    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    invoice = items.first()
    invoice_date = date_format(invoice.created_at, 'd.m.Y')

    ws.merge_cells('A1:F1')
    ws['A1'] = f"НАКЛАДНАЯ № {invoice.id}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center

    ws.merge_cells('A2:F2')
    ws['A2'] = f'от "{invoice_date}"'
    ws['A2'].alignment = Alignment(horizontal="right")

    ws['A4'] = "От кого:"
    ws['A5'] = f"Кому:"
    ws['B4'] = "OOO RATTAN MASTER"
    ws['B5'] = f"{invoice.product_to or ''}"

    headers = ["№ п/п", "Наименование", "Размер", "Цвет", "Ед. изм.", "Количество"]
    ws.append([])
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 20 if col_num != 1 else 8

    row = 8
    for index, item in enumerate(items, start=1):
        values = [
            index,
            item.name.title,
            item.size.title if item.size else '',
            item.color.title if item.color else '',
            "шт",
            item.quantity
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = center
            cell.border = border
        row += 1

    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Сдал: ________________   Ф. И. О."
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = "Принял: ________________   Ф. И. О."

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="InvoicePaper.xlsx"'
    wb.save(response)
    return response


@role_required(['Оператор склада Келес', 'Начальник'])
def invoice_keles_update(request, pk):
    invoice = get_object_or_404(InvoiceKeles, pk=pk)
    if request.method == 'POST':
        form = InvoiceKelesForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            return redirect('keles:invoice-list-keles')
    else:
        form = InvoiceKelesForm(instance=invoice)
    return render(request, 'keles-add/invoice_edit.html', {'form': form, 'invoice': invoice})


@role_required(['Оператор склада Келес', 'Начальник'])
def invoice_keles_delete(request, pk):
    invoice = get_object_or_404(InvoiceKeles, pk=pk)
    invoice.delete()
    return redirect('keles:invoice-list-keles')


@role_required(['Оператор склада', 'Начальник'])
def invoice_xasanboy_edit(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        form = InvoiceXasanboyForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            return redirect('pages:invoice-list')  # Или на detail: 'pages:invoice-detail', invoice.pk
    else:
        form = InvoiceKelesForm(instance=invoice)
    return render(request, 'edit/INV.html', {'form': form, 'invoice': invoice})


@role_required(['Оператор склада', 'Начальник'])
def invoice_xasanboy_delete(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    invoice.delete()
    return redirect('pages:invoice-list')
