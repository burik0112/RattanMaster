from datetime import datetime
from collections import defaultdict
from decimal import Decimal
from itertools import chain

import openpyxl
import xlwt
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.views import LoginView
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.utils.formats import date_format
from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from admiin.decorators import role_required
from admiin.forms import LoginForm
from keles.models import ProductEntryKeles, RemaingInventoryKeles, InvoiceCreateKeles
from .models import CategoryModel, ColorModel, SizeModel, InvoiceCreateModel, RemaingInventoryModel, ProductEntry, \
    TransferToInventory, ProductPriceModel, Invoice
from django.db.models import Count, Sum, Q


@role_required(['Начальник'])
def IndexCustom(request):
    total_1 = InvoiceCreateModel.objects.aggregate(total=Sum('quantity'))['total'] or 0
    total_2 = InvoiceCreateKeles.objects.aggregate(total=Sum('quantity'))['total'] or 0
    total_quantity = total_1 + total_2

    total_1 = ProductEntry.objects.aggregate(total=Sum('quantity'))['total'] or 0
    total_2 = ProductEntryKeles.objects.aggregate(total=Sum('quantity'))['total'] or 0
    total_product_in = total_1 + total_2
    total_remain_1 = RemaingInventoryModel.objects.aggregate(total=Sum('quantity'))['total'] or 0
    total_remain_2 = RemaingInventoryKeles.objects.aggregate(total=Sum('quantity'))['total'] or 0
    total_remaing = total_remain_1 + total_remain_2
    # Если нет данных, установим значение в 0
    if total_quantity is None:
        total_quantity = 0

    if total_product_in is None:
        total_product_in = 0

    if total_remaing is None:
        total_product_in = 0

    # Аннотируем категории, добавляя количество продуктов в каждой категории
    categories_with_product_count = CategoryModel.objects.annotate(
        count=Count('products'),
        count_keles=Count('products_keles')
    )


    # Возвращаем результат в шаблон
    return render(request, 'catalog_category.html', {
        'query': categories_with_product_count,
        'total_quantity': total_quantity,
        'total_product_in': total_product_in,
        'total_remaing': total_remaing,
    })


def Base(request, pk):
    query = InvoiceCreateModel.objects.filter(name__id=pk)
    return render(request, 'index.html', {'query': query})


def Detail(request, pk):
    query = InvoiceCreateModel.objects.filter(pk=pk)
    return render(request, 'detail.html', {'query': query})


@role_required(['Оператор склада', 'Начальник'])
def Categories(request):
    query = CategoryModel.objects.all()
    return render(request, 'category-cat.html', {'query': query})


@role_required(['Оператор склада', 'Начальник'])
def ColorAdd(request):
    query = ColorModel.objects.all()
    return render(request, 'color.html', {'query': query})


@role_required(['Оператор склада', 'Начальник'])
def Size(request):
    query = SizeModel.objects.all()  # Now this refers to the model
    return render(request, 'size-catalog.html', {'query': query})


@role_required(['Оператор склада', 'Начальник'])
def InvoiceCreate(request):
    invoice = InvoiceCreateModel.objects.all().order_by('-created_at')

    # Get filter values from the GET request
    product_id = request.GET.get('product_id')
    product_name = request.GET.get('product_name')
    product_size = request.GET.get('product_size')
    color = request.GET.get('product_color')
    transfer = request.GET.get('transfer')
    quantity = request.GET.get('quantity')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Apply filters based on the request
    if product_id:
        invoice = invoice.filter(name__id=product_id)
    if product_name:
        invoice = invoice.filter(name__title__icontains=product_name)
    if product_size:
        invoice = invoice.filter(size__title__icontains=product_size)
    if color:
        invoice = invoice.filter(color__title=color)
    if transfer:
        invoice = invoice.filter(product_to__title__icontains=transfer)
    if quantity:
        invoice = invoice.filter(quantity=quantity)

    # Apply date filter if provided
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            invoice = invoice.filter(created_at__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass  # Ignore invalid date formats

    # If export is requested, generate Excel file
    if request.GET.get('export') == 'excel':
        return export_to_excel(invoice)  # Call the export function

    # Render the template with filtered data
    return render(request, 'invoice-catalog.html', {'invoice': invoice})


@role_required(['Оператор склада', 'Начальник'])
def export_to_excel(invoices):
    # Create a new workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    # Define the header row
    headers = [
        'Invoice ID', 'Product Name', 'Size', 'Color', 'Transfer To', 'Quantity', 'Created At'
    ]
    ws.append(headers)

    # Populate data rows from the filtered invoices
    for inv in invoices:
        row = [
            str(inv.id),
            inv.name.title if inv.name else 'N/A',
            inv.size.title if inv.size else 'N/A',
            inv.color.title if inv.color else 'N/A',
            inv.product_to.title if inv.product_to else 'N/A',
            str(inv.quantity),
            inv.created_at.strftime('%Y-%m-%d')  # Format the date
        ]
        ws.append(row)

    # Create an HTTP response with content-type set for Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="InvoiceData.xlsx"'

    # Save the workbook directly to the response object
    wb.save(response)

    return response


@role_required(['Сотрудник приемки', 'Начальник'])
def ProductIn(request):
    # Start with getting all ProductEntry objects, ordered by created_at
    product_in = ProductEntry.objects.all().order_by('-created_at')

    # Get the search filters from the GET request
    search_query = request.GET.get('search', '')
    product_id = request.GET.get('product_id')
    product_name = request.GET.get('product_name')
    product_size = request.GET.get('product_size')
    color = request.GET.get('color')
    transfer = request.GET.get('transfer')
    quantity = request.GET.get('quantity')
    created_at = request.GET.get('created_at')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Apply the filters based on the search parameters
    if product_id:
        product_in = product_in.filter(name__id=product_id)
    if product_name:
        product_in = product_in.filter(name__title__icontains=product_name)
    if product_size:
        product_in = product_in.filter(size__title__icontains=product_size)
    if color:
        product_in = product_in.filter(color__title__icontains=color)
    if transfer:
        product_in = product_in.filter(product_in__title__icontains=transfer)
    if quantity:
        product_in = product_in.filter(quantity=quantity)

    # Apply the date filter if the 'created_at' filter is provided
    if created_at:
        try:
            # Convert the 'created_at' string into a date object
            created_at_date = datetime.strptime(created_at, '%Y-%m-%d').date()
            product_in = product_in.filter(created_at=created_at_date)
        except ValueError:
            pass  # If the date format is invalid, simply ignore it

    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            product_in = product_in.filter(created_at__range=[start_date_obj, end_date_obj])
        except ValueError:
            pass  # Ignore invalid date formats

    # If export is requested, generate Excel file
    if request.GET.get('export') == 'excel':
        return export_from_excel(product_in)

    return render(request, 'product_in-catalog.html', {
        'product_in': product_in,
        'search_query': search_query,
    })


@role_required(['Сотрудник приемки', 'Начальник'])
def export_from_excel(invoices):
    # Create a new workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    # Define the header row
    headers = [
        'Invoice ID', 'Product Name', 'Size', 'Color', 'TransferFrom', 'Quantity', 'Created At'
    ]
    ws.append(headers)

    # Populate data rows from the filtered invoices
    for inv in invoices:
        row = [
            str(inv.id),
            inv.name.title if inv.name else 'N/A',
            inv.size.title if inv.size else 'N/A',
            inv.color.title if inv.color else 'N/A',
            inv.product_in.title if inv.product_in else 'N/A',
            str(inv.quantity),
            inv.created_at.strftime('%Y-%m-%d')  # Format the date
        ]
        ws.append(row)

    # Create an HTTP response with content-type set for Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="InvoiceData.xlsx"'

    # Save the workbook directly to the response object
    wb.save(response)

    return response

@role_required(['Сотрудник приемки', 'Начальник'])
def RemaingList(request):
    remaing = RemaingInventoryModel.objects.all().order_by('-name')
    return render(request, 'remaing-list.html', {'remaing': remaing})


@role_required(['Менеджер склада', 'Начальник'])
def inventory_report(request):
    # Collect filters
    search_name = request.GET.get('search_name', '').strip().lower()
    search_id = request.GET.get('search_id', '').strip().lower()
    search_size = request.GET.get('search_size', '').strip().lower()
    search_color = request.GET.get('search_color', '').strip().lower()

    # Preload all product info (to avoid querying in loop)
    all_entries = list(ProductEntry.objects.select_related('name', 'size', 'color').all())
    all_remaining = list(RemaingInventoryModel.objects.select_related('name', 'size', 'color').all())
    all_invoice = list(InvoiceCreateModel.objects.select_related('name', 'size', 'color').all())

    # Aggregate quantities by (name_id, size_id, color_id)
    def aggregate_by(items):
        result = defaultdict(int)
        for obj in items:
            key = (obj.name_id, obj.size_id, obj.color_id)
            result[key] += obj.quantity
        return result

    entry_totals = aggregate_by(all_entries)
    remaining_totals = aggregate_by(all_remaining)
    invoice_totals = aggregate_by(all_invoice)

    # Combine all keys
    all_keys = set(entry_totals) | set(remaining_totals) | set(invoice_totals)

    inventory_data = []

    # Build a single reference map for title resolution
    ref_objects = {}
    for obj in all_entries + all_remaining + all_invoice:
        key = (obj.name_id, obj.size_id, obj.color_id)
        if key not in ref_objects:
            ref_objects[key] = obj

    for key in all_keys:
        name_id, size_id, color_id = key
        total_in = entry_totals.get(key, 0)
        total_remaining = remaining_totals.get(key, 0)
        total_invoice = invoice_totals.get(key, 0)
        remaining_stock = total_in + total_remaining - total_invoice

        obj = ref_objects.get(key)
        if not obj:
            continue

        category_title = obj.name.title.lower() if obj.name else ''
        size_title = obj.size.title.lower() if obj.size else ''
        color_title = obj.color.title.lower() if obj.color else ''
        category_id = str(obj.name.id) if obj.name else ''

        if (
                search_name in category_title and
                search_id in category_id and
                search_size in size_title and
                search_color in color_title
        ):
            inventory_data.append({
                'category_id': obj.name.id,
                'category_title': obj.name.title,
                'size_title': obj.size.title,
                'color_title': obj.color.title,
                'total_in': total_in,
                'total_remaining': total_remaining,
                'total_invoice': total_invoice,
                'remaining_stock': remaining_stock,
            })
    inventory_data.sort(key=lambda x: (
        x['category_title'].lower(),
        x['size_title'].lower(),
        x['color_title'].lower()
    ))

    # Excel download
    if 'download_excel' in request.GET:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"

        headers = ['ID', 'Category', 'Size', 'Color', 'Remaining']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header

        for row_num, data in enumerate(inventory_data, 2):
            ws[f'A{row_num}'] = data['category_id']
            ws[f'B{row_num}'] = data['category_title']
            ws[f'C{row_num}'] = data['size_title']
            ws[f'D{row_num}'] = data['color_title']
            ws[f'E{row_num}'] = data['remaining_stock']

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Turnover.xlsx"'
        wb.save(response)
        return response

    return render(request, 'turnover.html', {'inventory_data': inventory_data})

@role_required(['Менеджер склада', 'Начальник'])
def Dashboard(request, pk):
    # Fetching the data from the database
    query = InvoiceCreateModel.objects.filter(product_to__id=pk)

    if request.GET.get('export') == 'excel':  # Check if the export parameter is set
        # Create a workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Data"

        # Define the header row (you can customize this as per your model fields)
        headers = ['ID', 'Product Name', 'Color', 'Size', 'Product To', 'Quantity',
                   'Created_at']  # Replace with your model fields
        ws.append(headers)

        # Populate the data rows from the queryset
        for invoice in query:
            row = [
                str(invoice.id),  # Invoice ID
                str(invoice.name) if invoice.name else 'N/A',  # CategoryModel name (using .name)
                str(invoice.color) if invoice.color else 'N/A',  # ColorModel color (using .color)
                str(invoice.size) if invoice.size else 'N/A',
                str(invoice.product_to.title) if invoice.product_to else 'N/A',
                # TransferToInventory name (using .name)
                str(invoice.quantity),  # Quantity
                invoice.created_at.strftime('%Y-%m-%d'),  # Date formatted as YYYY-MM-DD
            ]
            row = [str(cell) for cell in row]
            ws.append(row)

            # Create an HTTP response with content-type set for Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Client List.xlsx'

        # Save the workbook directly to the response object
        wb.save(response)

        return response

        # If no export requested, render the normal dashboard view
    return render(request, 'dashboard.html', {'query': query})
    # SizeModel size (using .size)


def group_by_keys(queryset):
    data = defaultdict(int)
    for row in queryset:
        key = (row['name_id'], row['size_id'], row['color_id'])
        data[key] += row['quantity'] or 0
    return data


@role_required(['Менеджер склада', 'Начальник', 'Наблюдатель'])
def combined_inventory(request):
    # Search filters
    search_id = request.GET.get('search_id', '').strip().lower()
    search_name = request.GET.get('search_name', '').strip().lower()
    search_size = request.GET.get('search_size', '').strip().lower()
    search_color = request.GET.get('search_color', '').strip().lower()

    # Grouped data from all sources
    product_1_data = group_by_keys(
        ProductEntryKeles.objects.values('name_id', 'size_id', 'color_id').annotate(quantity=Sum('quantity')))
    remain_1_data = group_by_keys(
        RemaingInventoryKeles.objects.values('name_id', 'size_id', 'color_id').annotate(quantity=Sum('quantity')))
    invoice_1_data = group_by_keys(
        InvoiceCreateKeles.objects.values('name_id', 'size_id', 'color_id').annotate(quantity=Sum('quantity')))

    product_2_data = group_by_keys(
        ProductEntry.objects.values('name_id', 'size_id', 'color_id').annotate(quantity=Sum('quantity')))
    remain_2_data = group_by_keys(
        RemaingInventoryModel.objects.values('name_id', 'size_id', 'color_id').annotate(quantity=Sum('quantity')))
    invoice_2_data = group_by_keys(
        InvoiceCreateModel.objects.values('name_id', 'size_id', 'color_id').annotate(quantity=Sum('quantity')))

    # Build combined key set
    combined_keys = set(
        product_1_data.keys() | remain_1_data.keys() | invoice_1_data.keys() |
        product_2_data.keys() | remain_2_data.keys() | invoice_2_data.keys()
    )

    # Create a fast object lookup map
    object_map = {}
    all_objects = chain(
        ProductEntry.objects.select_related('name', 'size', 'color').all(),
        RemaingInventoryModel.objects.select_related('name', 'size', 'color').all(),
        InvoiceCreateModel.objects.select_related('name', 'size', 'color').all(),
        ProductEntryKeles.objects.select_related('name', 'size', 'color').all(),
        RemaingInventoryKeles.objects.select_related('name', 'size', 'color').all(),
        InvoiceCreateKeles.objects.select_related('name', 'size', 'color').all(),
    )

    for obj in all_objects:
        key = (obj.name_id, obj.size_id, obj.color_id)
        if key not in object_map:
            object_map[key] = obj

    # Build inventory data
    inventory_data = []
    for key in combined_keys:
        name_id, size_id, color_id = key

        remaining_1 = product_1_data.get(key, 0) + remain_1_data.get(key, 0) - invoice_1_data.get(key, 0)
        remaining_2 = product_2_data.get(key, 0) + remain_2_data.get(key, 0) - invoice_2_data.get(key, 0)

        obj = object_map.get(key)
        if obj:
            category_id = str(obj.name.id)
            category_title = obj.name.title.lower()
            size_title = obj.size.title.lower()
            color_title = obj.color.title.lower()

            if (
                    search_id in category_id and
                    search_name in category_title and
                    search_size in size_title and
                    search_color in color_title
            ):
                inventory_data.append({
                    'category_id': obj.name.id,
                    'category_title': obj.name.title,
                    'size_title': obj.size.title,
                    'color_title': obj.color.title,
                    'remaining_2': remaining_2,
                    'remaining_1': remaining_1,
                    'total_remaining': remaining_1 + remaining_2
                })
    inventory_data.sort(key=lambda x: (
        x['category_title'].lower(),
        x['size_title'].lower(),
        x['color_title'].lower()
    ))

    if 'download_excel' in request.GET:
        if request.user.role not in ['Менеджер склада', 'Начальник']:
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventory Report"

            headers = ['ID', 'Category', 'Size', 'Color', 'Xasanboy', 'Keles', 'Total Remaining']
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'] = header

            for row_num, data in enumerate(inventory_data, 2):
                ws[f'A{row_num}'] = data['category_id']
                ws[f'B{row_num}'] = data['category_title']
                ws[f'C{row_num}'] = data['size_title']
                ws[f'D{row_num}'] = data['color_title']
                ws[f'E{row_num}'] = data['remaining_2']
                ws[f'F{row_num}'] = data['remaining_1']
                ws[f'G{row_num}'] = data['total_remaining']

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="TotalTurnover.xlsx"'
            wb.save(response)
            return response

    # Return rendered template
    return render(request, 'totalturnover.html', {
        'inventory_data': inventory_data
    })


def shop_list(request):
    shops = TransferToInventory.objects.all()

    return render(request, 'transfer_to-list.html', {'shops': shops})


def shop_summary(request, pk):
    # Get the shop by pk
    shop = get_object_or_404(TransferToInventory, pk=pk)

    invoices = InvoiceCreateModel.objects.filter(
        invoice__product_to=shop
    ).select_related('name', 'size', 'color', 'invoice__product_to')

    summary = defaultdict(lambda: {
        'quantity': 0,
        'total': Decimal('0.00'),
        'name': '',
        'size': '',
        'color': '',
        'price': Decimal('0.00')  # <-- Har bir guruh uchun narxni saqlaymiz
    })

    for inv in invoices:
        key = (inv.name.id, inv.size.id, inv.color.id if inv.color else None)

        try:
            price_obj = ProductPriceModel.objects.get(
                name=inv.name,
                size=inv.size,
                color=inv.color
            )
            price = price_obj.price
        except ProductPriceModel.DoesNotExist:
            price = Decimal('0.00')

        summary[key]['quantity'] += inv.quantity
        summary[key]['total'] += price * inv.quantity
        summary[key]['name'] = inv.name.title
        summary[key]['size'] = inv.size.title
        summary[key]['color'] = inv.color.title if inv.color else '-'
        summary[key]['price'] = price  # <-- Narxni shu yerda saqlaymiz

    aggregated_data = list(summary.values())

    total_quantity = sum(item['quantity'] for item in aggregated_data)
    total_amount = sum(item['total'] for item in aggregated_data)

    return render(request, 'shop_report.html', {
        'aggregated_data': aggregated_data,
        'shop': shop,
        'total_quantity': total_quantity,
        'total_amount': total_amount
    })


def shop_export_excel(request, pk):
    shop = get_object_or_404(TransferToInventory, pk=pk)
    invoices = InvoiceCreateModel.objects.filter(product_to=shop).select_related('name', 'size', 'color')

    summary = defaultdict(lambda: {'quantity': 0, 'total': Decimal('0.00'), 'name': '', 'size': '', 'color': '',
                                   'price': Decimal('0.00')})
    for inv in invoices:
        key = (inv.name.id, inv.size.id, inv.color.id if inv.color else None)

        try:
            price_obj = ProductPriceModel.objects.get(name=inv.name, size=inv.size, color=inv.color)
            price = price_obj.price
        except ProductPriceModel.DoesNotExist:
            price = Decimal('0.00')

        summary[key]['quantity'] += inv.quantity
        summary[key]['total'] += price * inv.quantity
        summary[key]['name'] = inv.name.title
        summary[key]['size'] = inv.size.title
        summary[key]['color'] = inv.color.title if inv.color else '-'
        summary[key]['price'] = price

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{shop.title}_products.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Products')

    columns = ['Product Name', 'Size', 'Color', 'Quantity', 'Price per1', 'Total Amount']
    for col_num in range(len(columns)):
        ws.write(0, col_num, columns[col_num])

    row = 1
    for item in summary.values():
        ws.write(row, 0, item['name'])
        ws.write(row, 1, item['size'])
        ws.write(row, 2, item['color'])
        ws.write(row, 3, item['quantity'])
        ws.write(row, 4, float(item['price']))
        ws.write(row, 5, float(item['total']))
        row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="TotalSum.xlsx"'
    wb.save(response)
    return response


def login_view(request):
    form = LoginForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        username = form.cleaned_data['username']
        password = form.cleaned_data['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            role = getattr(user, 'role', None)
            if role == 'Начальник':
                return redirect('pages:cards')
            elif role == 'Менеджер склада':
                return redirect('pages:turnover_list')
            elif role == 'Оператор склада':
                return redirect('pages:invoice-list')
            elif role == 'Сотрудник приемки':
                return redirect('pages:product_in-list')
            elif role == 'Менеджер склада Келес':
                return redirect('keles:turnover_list')
            elif role == 'Оператор склада Келес':
                return redirect('keles:invoice-list')
            elif role == 'Сотрудник приемки Келес':
                return redirect('keles:product_in-list')
            elif role == 'Наблюдатель':
                return redirect('pages:total-turnover_list')

    return render(request, 'registration/login.html', {'form': form})



def logout_view(request):
    logout(request)
    return redirect('pages:logout-success')


def logout_success_view(request):
    return render(request, 'registration/logout_success.html')

@role_required(['Менеджер склада', 'Начальник', 'Оператор склада'])
def invoice_list(request):
    # Извлекаем все накладные
    invoices = Invoice.objects.all().order_by('-created_at')  # Сортируем по дате создания (по убыванию)

    # Передаем данные в контекст
    context = {
        'invoices': invoices
    }

    return render(request, 'invoice_list.html', context)


@role_required(['Менеджер склада', 'Начальник', 'Оператор склада'])
def invoice_detail(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    items = invoice.items.all()  # через related_name='items'

    context = {
        'invoice': invoice,
        'items': items
    }
    return render(request, 'invoice_detail.html', context)

@role_required(['Оператор склада'])
def export_to_excel(request, invoice_id):
    try:
        invoice = Invoice.objects.get(pk=invoice_id)
    except Invoice.DoesNotExist:
        return HttpResponse("Накладная не найдена.")

    items = invoice.items.select_related('name', 'size', 'color')
    if not items.exists():
        return HttpResponse("Нет товаров для экспорта.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Накладная"

    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    invoice_date = date_format(invoice.created_at, 'd.m.Y')

    # Шапка
    ws.merge_cells('A1:F1')
    ws['A1'] = f"НАКЛАДНАЯ № {invoice.number}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center

    ws.merge_cells('A2:F2')
    ws['A2'] = f'от "{invoice_date}"'
    ws['A2'].alignment = center

    ws['A4'] = "От кого:"
    ws['B4'] = "OOO RATTAN MASTER"
    ws['A5'] = "Кому:"
    ws['B5'] = invoice.product_to.title if invoice.product_to else "—"
    ws['D5'] = "Через________________"

    # Таблица
    headers = ["№ п/п", "Наименование", "Размер", "Цвет", "Количество"]
    ws.append([])
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 20 if col_num != 1 else 8

    row = 8
    total_quantity = 0

    for index, item in enumerate(items, start=1):
        values = [
            index,
            item.name.title,
            item.size.title if item.size else '',
            item.color.title if item.color else '',
            f"{item.quantity} шт"
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = center
            cell.border = border
        total_quantity += item.quantity
        row += 1

    # Итого
    total_row = ["", "", "", "Итого:", f"{total_quantity} шт"]
    for col, val in enumerate(total_row, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = bold_font
        cell.alignment = center
        cell.border = border
    row += 2

    # Подписи
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "Сдал: ________________   Ф. И. О."
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'] = "Принял: ________________   Ф. И. О."

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Invoice_{invoice.number}.xlsx"'
    wb.save(response)
    return response